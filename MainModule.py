import sys
from PyQt5 import QtWidgets, QtCore, Qt
from PyQt5.QtWidgets import QMessageBox, QErrorMessage
from math import ceil
from PyQt5.QtGui import QIcon, QFont
from PyQt5.QtCore import QSize, QObject
from PyQt5.QtWidgets import qApp, QInputDialog, QFileDialog, QPushButton, QTableWidgetItem, QHeaderView
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side


import Interface_main
import Instructions
import ModuleNanoSubstances


class InstructionsClass(QtWidgets.QWidget, Instructions.Ui_Form_instructions):
    def __init__(self):
        super().__init__()
        self.setupUi(self)


class MainClass(QtWidgets.QMainWindow, Interface_main.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.Interface_modification()

        self.kontrol_analiz_common = False
        self.kontrol_analiz_glass = False

        self.Calc_koef_ltransm.clicked.connect(self.Calc_koef_ltransm_glass)
        self.Give_conclusion.clicked.connect(self.Give_conclusion_glass)
        self.Def_security_category.clicked.connect(self.Give_Hazard_assessment_common)

        self.action_quit.triggered.connect(qApp.quit)
        self.action_Excel_common.triggered.connect(self.createExcelReport_common)
        self.action_Excel_glass.triggered.connect(self.createExcelReport_glass)

        self.action_about.triggered.connect(self.AboutClicked)
        self.action_help.triggered.connect(self.HelpClicked)

    def Interface_modification(self):
        self.Icon = QIcon("logo\\icon.ico")
        self.setWindowIcon(self.Icon)

        self.table_1.setColumnWidth(0, 115)
        self.table_1.setColumnWidth(1, 130)
        self.table_1.setColumnWidth(2, 125)
        self.table_1.setColumnWidth(3, 170)
        self.table_1.setColumnWidth(4, 39)

        self.table_2.setColumnWidth(0, 115)
        self.table_2.setColumnWidth(1, 130)
        self.table_2.setColumnWidth(2, 125)
        self.table_2.setColumnWidth(3, 170)
        self.table_2.setColumnWidth(4, 39)

        self.but_delete_nano_1 = []
        self.but_add_nano_1 = QPushButton(self)
        self.but_add_nano_1.setStyleSheet("background-color: #FFFFFF;")
        self.but_add_nano_1.setIcon(QIcon('logo/plus.png'))
        self.but_add_nano_1.setIconSize(QSize(20, 20))
        self.table_1.setCellWidget(0, 4, self.but_add_nano_1)
        self.but_add_nano_1.clicked.connect(self.create_nano_substances_common)

        self.but_delete_nano_2 = []
        self.but_add_nano_2 = QPushButton(self)
        self.but_add_nano_2.setStyleSheet("background-color: #FFFFFF;")
        self.but_add_nano_2.setIcon(QIcon('logo/plus.png'))
        self.but_add_nano_2.setIconSize(QSize(20, 20))
        self.table_2.setCellWidget(0, 4, self.but_add_nano_2)
        self.but_add_nano_2.clicked.connect(self.create_nano_substances_glass)

    def create_nano_substances_common(self):
        window_nano_substances = ModuleNanoSubstances.ClassNanoSubstances(self)
        window_nano_substances.signal_close.connect(self.add_nano_substances_common)
        window_nano_substances.show()

    def create_nano_substances_glass(self):
        window_nano_substances = ModuleNanoSubstances.ClassNanoSubstances(self)
        window_nano_substances.signal_close.connect(self.add_nano_substances_glass)
        window_nano_substances.show()

    def gen_name_nano_substances(self, name_sub, table):
        rowPosition = table.rowCount() - 1
        for i in range(rowPosition):
            name = table.verticalHeaderItem(i).text()
            if name == name_sub:
                return self.gen_name_nano_substances(name_sub + '_', table)
        return name_sub

    def add_nano_substances_common(self, struktura_data):
        name = self.gen_name_nano_substances(struktura_data[0], self.table_1)

        rowPosition = self.table_1.rowCount() - 1
        self.table_1.insertRow(rowPosition)
        item = QTableWidgetItem(name)
        font = QFont()
        font.setPointSize(8)
        item.setFont(font)
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.table_1.setVerticalHeaderItem(rowPosition, item)
        for column in range(4):
            item = QTableWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            item.setFlags(QtCore.Qt.ItemIsSelectable |
                          QtCore.Qt.ItemIsDragEnabled | QtCore.Qt.ItemIsDropEnabled |
                          QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            self.table_1.setItem(rowPosition, column, item)
        self.table_1.setRowHeight(rowPosition, 50)

        self.table_1.item(rowPosition, 0).setText(str(struktura_data[1]))
        self.table_1.item(rowPosition, 1).setText(struktura_data[2])
        self.table_1.item(rowPosition, 2).setText(str(struktura_data[3]))
        self.table_1.item(rowPosition, 3).setText(struktura_data[4][0])
        self.table_1.item(rowPosition, 3).setToolTip(struktura_data[4][1])

        self.but_delete_nano_1.append(QPushButton(self))
        self.but_delete_nano_1[rowPosition].setStyleSheet("background-color: #FFFFFF;")
        self.but_delete_nano_1[rowPosition].setIcon(QIcon('logo/minus.png'))
        self.but_delete_nano_1[rowPosition].setIconSize(QSize(20, 20))
        self.but_delete_nano_1[rowPosition].clicked.connect(
            lambda checked, row_name=name: self.del_row_table_1(row_name)
        )
        self.table_1.setCellWidget(rowPosition, 4, self.but_delete_nano_1[rowPosition])

    def del_row_table_1(self, row_name):
        rowPosition = self.table_1.rowCount() - 1
        row = None
        for i in range(rowPosition):
            name = self.table_1.verticalHeaderItem(i).text()
            if name == row_name:
                row = i

        self.table_1.removeRow(row)
        self.but_delete_nano_1.pop(row)

    def add_nano_substances_glass(self, struktura_data):
        name = self.gen_name_nano_substances(struktura_data[0], self.table_2)

        rowPosition = self.table_2.rowCount() - 1
        self.table_2.insertRow(rowPosition)
        item = QTableWidgetItem(name)
        font = QFont()
        font.setPointSize(8)
        item.setFont(font)
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        self.table_2.setVerticalHeaderItem(rowPosition, item)
        for column in range(4):
            item = QTableWidgetItem()
            item.setTextAlignment(QtCore.Qt.AlignCenter)
            item.setFlags(QtCore.Qt.ItemIsSelectable |
                          QtCore.Qt.ItemIsDragEnabled | QtCore.Qt.ItemIsDropEnabled |
                          QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            self.table_2.setItem(rowPosition, column, item)
        self.table_2.setRowHeight(rowPosition, 50)

        self.table_2.item(rowPosition, 0).setText(str(struktura_data[1]))
        self.table_2.item(rowPosition, 1).setText(struktura_data[2])
        self.table_2.item(rowPosition, 2).setText(str(struktura_data[3]))
        self.table_2.item(rowPosition, 3).setText(struktura_data[4][0])
        self.table_2.item(rowPosition, 3).setToolTip(struktura_data[4][1])

        self.but_delete_nano_2.append(QPushButton(self))
        self.but_delete_nano_2[rowPosition].setStyleSheet("background-color: #FFFFFF;")
        self.but_delete_nano_2[rowPosition].setIcon(QIcon('logo/minus.png'))
        self.but_delete_nano_2[rowPosition].setIconSize(QSize(20, 20))
        self.but_delete_nano_2[rowPosition].clicked.connect(
            lambda checked, row_name=name: self.del_row_table_2(row_name)
        )
        self.table_2.setCellWidget(rowPosition, 4, self.but_delete_nano_2[rowPosition])

    def del_row_table_2(self, row_name):
        rowPosition = self.table_2.rowCount() - 1
        row = None
        for i in range(rowPosition):
            name = self.table_2.verticalHeaderItem(i).text()
            if name == row_name:
                row = i

        self.table_2.removeRow(row)
        self.but_delete_nano_2.pop(row)

    def Give_Hazard_assessment_common(self):
        try:
            self.f_grade_1.setText("")
            self.Safety_category_1.setText("")

            self.kontrol_analiz_common = False
            self.name_object_1 = self.Edit_name_obj_1.text()
            if len(self.name_object_1) == 0:
                QMessageBox.question(self, 'Небыли введены данные',
                                     "Вы не ввели название объекта исследования\n"
                                     "Пожалуйста введите название и повторите попытку",
                                     QMessageBox.Ok)
                return

            self.row_count_1 = self.table_1.rowCount() - 1
            if self.row_count_1 == 0:
                QMessageBox.question(self, 'Отсутствует анализ нановеществ',
                                     "Вы не добавили ни одного нановещества, входящего в состав объекта.\n"
                                     "Пожалуйста добавьте все нановещества, входящие в состов объекта,\n"
                                     "и повторите анализ!",
                                     QMessageBox.Ok)
                return

            # Определение общей степени потенциальной опасности исследуемого объекта
            Hazard_assessments = [0]
            for i in range(self.row_count_1):
                if float(self.table_1.item(i, 2).text()) <= 0.25:
                    Hazard_assessments.append(float(self.table_1.item(i, 0).text()))

            max_Hazard_assessments = max(Hazard_assessments)
            self.Hazard_assessment_1_str = ""
            if 0.441 <= max_Hazard_assessments <= 1.11:
                self.Hazard_assessment_1_str = "Низкая степень потенциальной опасности"
            elif 1.111 <= max_Hazard_assessments <= 1.779:
                self.Hazard_assessment_1_str = "Средняя степень потенциальной опасности"
            elif 1.78 <= max_Hazard_assessments <= 2.449:
                self.Hazard_assessment_1_str = "Высокая степень потенциальной опасности"
            else:
                QMessageBox.question(self, 'Ошибка при определении общей степени потенциальной опасности!',
                                     "Вы не добавили ни одного нановещества,\nс достоверной оценкой опасности.\n",
                                     QMessageBox.Ok)
                self.Hazard_assessment_1_str = "Потенциальная опасность неопределена"
            self.f_grade_1.setText(self.Hazard_assessment_1_str)

            # Определение категории безопасности
            if self.Hazard_assessment_1_str == "Низкая степень потенциальной опасности":
                self.Safety_category_1_str = 'Категория А: "безопасно"'
            elif self.Hazard_assessment_1_str == "Средняя степень потенциальной опасности":
                self.Safety_category_1_str = 'Категория В "требует дополнительных исследований"'
            else:
                self.Safety_category_1_str = 'Категория С "безопасность не доказана"'
            self.Safety_category_1.setText("<center>" + self.Safety_category_1_str + "</center>")

            self.kontrol_analiz_common = True
            return
        except Exception as ex:
            QMessageBox.question(self, 'Ошибка в программе',
                                 "Ошибка: " + str(type(ex)) + "; Значение: " + str(ex),
                                 QMessageBox.Ok)

    def Calc_koef_ltransm_glass(self):
        with_glass_str = self.Edit_ltransm_with_glass.text()
        if len(with_glass_str) == 0.0:
            QMessageBox.question(self, 'Небыли введены данные',
                                 "Вы не ввели cветопропускание со стеклом.\n"
                                 "Пожалуйста введите его и повторите попытку!",
                                 QMessageBox.Ok)
            return
        without_glass_str = self.Edit_ltransm_without_glass.text()
        if len(without_glass_str) == 0.0:
            QMessageBox.question(self, 'Небыли введены данные',
                                 "Вы не ввели cветопропускание без стекла.\n"
                                 "Пожалуйста введите его и повторите попытку!",
                                 QMessageBox.Ok)
            return
        try:
            koef_ltransm = ceil(float(with_glass_str) / float(without_glass_str) * 10000) / 100
            self.Edit_koef_ltransm.setText(str(koef_ltransm))
            return
        except:
            QMessageBox.question(self, 'Введены неверные данные',
                                 "Вы ввели вместо чисел строку, или нулевой светопопускание без стекла.\n"
                                 "Пожалуйста введите верные данные и повторите попытку!",
                                 QMessageBox.Ok)
            return

    def Give_conclusion_glass(self):
        try:
            self.Level_ltransm.setText("")
            self.Glass_compliance.setText("")
            self.Status.setText("")
            self.f_grade_2.setText("")
            self.Safety_category_2.setText("")
            self.Conclusion.setText("")

            self.kontrol_analiz_glass = False
            self.name_object_2 = self.Edit_name_obj_2.text()
            if len(self.name_object_2) == 0:
                QMessageBox.question(self, 'Небыли введены данные',
                                     "Вы не ввели название объекта исследования\n"
                                     "Пожалуйста введите название и повторите попытку",
                                     QMessageBox.Ok)
                return

            self.koef_ltransm_str = self.Edit_koef_ltransm.text()
            try:
                koef_ltransm = float(self.koef_ltransm_str)
            except:
                QMessageBox.question(self, 'Введены неверные данные',
                                     "Вы ввели в качестве коэффициента светопропускания строку.\n"
                                     "Пожалуйста введите верные данные и повторите попытку!",
                                     QMessageBox.Ok)
                return

            # Оперделение уровень светопропускания
            level_ltransm = 0
            self.level_ltransm_str = ""
            if koef_ltransm >= 75.0 and koef_ltransm <= 100.0:
                level_ltransm = 1
                self.level_ltransm_str = "1й уровень (отличное светопропускание)"
            elif koef_ltransm >= 50.0 and koef_ltransm < 75.0:
                level_ltransm = 2
                self.level_ltransm_str = "2й уровень (хорошее светопропускание)"
            elif koef_ltransm >= 25.0 and koef_ltransm < 50.0:
                level_ltransm = 3
                self.level_ltransm_str = "3й уровень (значительное затемнение)"
            elif koef_ltransm >= 0.0 and koef_ltransm < 25.0:
                level_ltransm = 4
                self.level_ltransm_str = "4й уровень (темно)"
            else:
                QMessageBox.question(self, 'Введены неверные данные',
                                     "Вы ввели коэффициента светопропускания меньше 0% или больше 100%.\n"
                                     "Пожалуйста введите верные данные и повторите попытку!",
                                     QMessageBox.Ok)
                return
            self.Level_ltransm.setText(self.level_ltransm_str)

            self.Glass_compliance_str = ""
            self.status_str = ""
            self.Conclusion_str = ""
            self.Safety_category_2_str = ""
            self.Glass_compliance.setText(self.Glass_compliance_str)
            self.Status.setText(self.status_str)
            self.Conclusion.setText(self.Conclusion_str)
            self.Safety_category_2.setText(self.Safety_category_2_str)

            # Оперделение допустимого соответствия и статуса стекла
            if self.Func_coverage.currentIndex() == 0:
                self.Glass_compliance_str = "Допущено"
                self.Glass_compliance.setText(self.Glass_compliance_str)
                if level_ltransm == 1 or level_ltransm == 2:
                    self.status_str = "Приоритетный"
                    self.Status.setText(self.status_str)
                else:
                    self.status_str = ""
                    self.Status.setText(self.status_str)
            elif self.Func_coverage.currentIndex() == 1 or self.Func_coverage.currentIndex() == 3:
                if level_ltransm == 1 or level_ltransm == 2:
                    self.Glass_compliance_str = "Допущено"
                    self.Glass_compliance.setText(self.Glass_compliance_str)
                    if level_ltransm == 1:
                        self.status_str = "Приоритетный"
                        self.Status.setText(self.status_str)
                    else:
                        self.status_str = ""
                        self.Status.setText(self.status_str)
                else:
                    self.Conclusion_str = "Рекомендация образцов с" \
                                         "применением наноматериалов не обоснована"
                    self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")
                    self.Conclusion.setText(self.Conclusion_str)
                    self.Safety_category_2_str = "—————"
                    self.Safety_category_2.setText("<center>" + self.Safety_category_2_str + "</center>")
                    self.kontrol_analiz_glass = True
                    return
            else:
                if level_ltransm == 2 or level_ltransm == 3:
                    self.Glass_compliance_str = "Допущено"
                    self.Glass_compliance.setText(self.Glass_compliance_str)
                    if level_ltransm == 2:
                        self.status_str = "Приоритетный"
                        self.Status.setText(self.status_str)
                    else:
                        self.status_str = ""
                        self.Status.setText(self.status_str)
                else:
                    self.Conclusion_str = "Рекомендация образцов с"\
                                          "применением наноматериалов не обоснована"
                    self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")
                    self.Safety_category_2_str = "—————"
                    self.Safety_category_2.setText("<center>" + self.Safety_category_2_str + "</center>")
                    self.kontrol_analiz_glass = True
                    return

            self.row_count_2 = self.table_2.rowCount() - 1
            if self.row_count_2 == 0:
                QMessageBox.question(self, 'Отсутствует анализ нановеществ',
                                     "Вы не добавили ни одного нановещества, входящего в состав объекта.\n"
                                     "Пожалуйста добавьте все нановещества, входящие в состов объекта,\n"
                                     "и повторите анализ!",
                                     QMessageBox.Ok)
                return

            # Определение общей степени потенциальной опасности исследуемого объекта
            Hazard_assessments = [0]
            for i in range(self.row_count_2):
                if float(self.table_2.item(i, 2).text()) <= 0.25:
                    Hazard_assessments.append(float(self.table_2.item(i, 0).text()))

            max_Hazard_assessments = max(Hazard_assessments)
            self.Hazard_assessment_2_str = ""
            if 0.441 <= max_Hazard_assessments <= 1.11:
                self.Hazard_assessment_2_str = "Низкая степень потенциальной опасности"
            elif 1.111 <= max_Hazard_assessments <= 1.779:
                self.Hazard_assessment_2_str = "Средняя степень потенциальной опасности"
            elif 1.78 <= max_Hazard_assessments <= 2.449:
                self.Hazard_assessment_2_str = "Высокая степень потенциальной опасности"
            else:
                QMessageBox.question(self, 'Ошибка при определении общей степени потенциальной опасности!',
                                     "Вы не добавили ни одного нановещества,\nс достоверной оценкой опасности.\n",
                                     QMessageBox.Ok)
                self.Hazard_assessment_2_str = "Потенциальная опасность неопределена"
            self.f_grade_2.setText(self.Hazard_assessment_2_str)


            # Определение категории безопасности
            if self.Hazard_assessment_2_str == "Низкая степень потенциальной опасности":
                self.Safety_category_2_str = 'Категория А: "безопасно"'
            elif self.Hazard_assessment_2_str == "Средняя степень потенциальной опасности":
                self.Safety_category_2_str = 'Категория В "требует дополнительных исследований"'
            else:
                self.Safety_category_2_str = 'Категория С "безопасность не доказана"'
            self.Safety_category_2.setText("<center>" + self.Safety_category_2_str + "</center>")

            # Финальное на основе статуса
            if self.status_str == "Приоритетный" and \
               self.Safety_category_2_str == 'Категория А: "безопасно"':
                self.Conclusion_str = "Рекомендовано к применению"
                self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")
            elif self.status_str == "Приоритетный" and \
                 self.Safety_category_2_str == 'Категория В "требует дополнительных исследований"':
                self.Conclusion_str = "Рекомендовано после положительного лабораторного " \
                                      "тестирования биохимической безопасности"
                self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")
            elif self.status_str == "Приоритетный" and \
                 self.Safety_category_2_str == 'Категория С "безопасность не доказана"':
                self.Conclusion_str = "Рекомендации ограничены. Требуется " \
                                      "расширенное биохимическое исследование"
                self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")

            # Финальное на основе допустимого соответствия
            elif self.Glass_compliance_str == "Допущено" and \
               self.Safety_category_2_str == 'Категория А: "безопасно"':
                self.Conclusion_str = "Рекомендовано к применению"
                self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")
            elif self.Glass_compliance_str == "Допущено" and \
                 self.Safety_category_2_str == 'Категория В "требует дополнительных исследований"':
                self.Conclusion_str = "Рекомендовано при положительных лабораторных " \
                                      "биохимических тестах и дополнительных тестах " \
                                      "эффективности целевой функции"
                self.Conclusion.setText('<center style="font-size:10pt;">' + self.Conclusion_str + '</center>')
            elif self.Glass_compliance_str == "Допущено" and \
                 self.Safety_category_2_str == 'Категория С "безопасность не доказана"':
                self.Conclusion_str = "Допущено при положительных результатах расширенного " \
                                      "биохимического исследования и дополнительных тестов " \
                                      "эффективности целевой функции"
                self.Conclusion.setText('<center style="font-size:10pt;">' + self.Conclusion_str + '</center>')
            else:
                self.Conclusion_str = "Рекомендация образцов с" \
                                      "применением наноматериалов не обоснована"
                self.Conclusion.setText("<center>" + self.Conclusion_str + "</center>")

            self.kontrol_analiz_glass = True
            return
        except Exception as ex:
            QMessageBox.question(self, 'Ошибка в программе',
                                 "Ошибка: " + str(type(ex)) + "; Значение: " + str(ex),
                                 QMessageBox.Ok)

    def Init_Excel_style(self):
        self.font_8 = Font(name='Times New Roman',
                         size=8,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000')

        self.font_11 = Font(name='Times New Roman',
                         size=11,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000')

        self.fill = PatternFill(fill_type='solid',
                                start_color='c1c1c1',
                                end_color='c2c2c2')

        self.border = Border(left=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             top=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             diagonal=Side(border_style='thin', color='FF000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin', color='FF000000'),
                             vertical=Side(border_style='thin', color='FF000000'),
                             horizontal=Side(border_style='thin', color='FF000000'))

        self.align_center = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=True,
                                      shrink_to_fit=False,
                                      indent=0)

        self.align_left = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=True,
                                    shrink_to_fit=False,
                                    indent=0)

        self.align_right = Alignment(horizontal='right',
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

    def createExcelReport_common(self):
        try:
            if not self.kontrol_analiz_common:
                QMessageBox.question(self, 'Ошибка сохранения',
                                     "Сохранение не удалось, так\n"
                                     "как не был проведён анализ данных!",
                                     QMessageBox.Ok)
                return

            file_path = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Электронная таблица (*.xlsx)")[0]
            self.Init_Excel_style()
            wb = Workbook()
            ws = wb.active
            ws.title = self.name_object_1

            for cellObj in ws['A1:A5']:
                for cell in cellObj:
                    ws[cell.coordinate].font = self.font_11

            end_table = 1 + self.row_count_1 + 2
            for cellObj in ws['A2:E'+str(end_table)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = self.font_8
                    ws[cell.coordinate].alignment = self.align_center

            end_doc = end_table + 2
            for cellObj in ws['A'+str(end_table+1)+':E'+str(end_doc)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = self.font_11

            for cellObj in ws['A1:E'+str(end_doc)]:
                for cell in cellObj:
                    ws[cell.coordinate].border = self.border

            ws.merge_cells('A1:B1')
            ws.merge_cells('C1:E1')
            ws['A1'].alignment = self.align_left
            ws['C1'].alignment = self.align_center

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18

            ws['A1'] = "Обозначение объекта:"

            ws['C1'] = self.name_object_1

            ws['A2'] = "Название нановещества"
            ws['B2'] = "Количественная оценка опасности"
            ws['C2'] = "Качественная оценка опасности"
            ws['D2'] = "Коэффициент неполноты анализа"
            ws['E2'] = "Оценка достоверности результатов"

            alf = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E'}
            for i in range(self.row_count_1):
                k = 1
                ws[alf[k] + str(i + 3)] = self.table_1.verticalHeaderItem(i).text()
                for j in range(4):
                    k += 1
                    ws[alf[k] + str(i + 3)] = self.table_1.item(i, j).text()

            ws.merge_cells('A'+str(end_table+1)+':B'+str(end_table+1))
            ws.merge_cells('C'+str(end_table+1)+':E'+str(end_table+1))
            ws.merge_cells('A'+str(end_table+2)+':B'+str(end_table+2))
            ws.merge_cells('C'+str(end_table+2)+':E'+str(end_table+2))
            ws['A'+str(end_table+1)].alignment = self.align_left
            ws['C'+str(end_table+1)].alignment = self.align_center
            ws['A'+str(end_table+2)].alignment = self.align_left
            ws['C'+str(end_table+2)].alignment = self.align_center
            ws.row_dimensions[end_table+1].height = 30

            ws['A'+str(end_table+1)] = "Общая степень потенциальной опасности исследуемого объекта:"
            ws['C'+str(end_table+1)] = self.Hazard_assessment_1_str
            ws['A'+str(end_table+2)] = "Категория безопасности:"
            ws['C'+str(end_table+2)] = self.Safety_category_1_str

            wb.save(file_path)
            return
        except PermissionError:
            QMessageBox.question(self, 'Ошибка сохранения',
                                 "Сохранение не удалось, проверьте не\n"
                                 "открыт ли файл в другой программе",
                                 QMessageBox.Ok)
        except FileNotFoundError:
            QMessageBox.question(self, 'Ошибка сохранения',
                                 "Вы не задали имя файла!",
                                 QMessageBox.Ok)
        except Exception as ex:
            QMessageBox.question(self, 'Ошибка в программе',
                                 "Ошибка: " + str(type(ex)) + "; Значение: " + str(ex),
                                 QMessageBox.Ok)

    def createExcelReport_glass(self):
        try:
            if not self.kontrol_analiz_glass:
                QMessageBox.question(self, 'Ошибка сохранения',
                                     "Сохранение не удалось, так\n"
                                     "как не был проведён анализ данных!",
                                     QMessageBox.Ok)
                return

            file_path = QFileDialog.getSaveFileName(self, "Сохранить файл", "", "Электронная таблица (*.xlsx)")[0]
            self.Init_Excel_style()
            wb = Workbook()
            ws = wb.active
            ws.title = self.name_object_2

            for cellObj in ws['A1:E5']:
                for cell in cellObj:
                    ws[cell.coordinate].font = self.font_11

            end_table = 5 + self.row_count_2 + 2
            for cellObj in ws['A6:E'+str(end_table)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = self.font_8
                    ws[cell.coordinate].alignment = self.align_center

            end_doc = end_table + 4
            for cellObj in ws['A'+str(end_table+1)+':E'+str(end_doc)]:
                for cell in cellObj:
                    ws[cell.coordinate].font = self.font_11

            for cellObj in ws['A1:E'+str(end_doc)]:
                for cell in cellObj:
                    ws[cell.coordinate].border = self.border

            for i in range(1, 6):
                ws.merge_cells('A'+str(i)+':B'+str(i))
                ws['A'+str(i)].alignment = self.align_left
                ws['C' + str(i)].alignment = self.align_center
                if i != 4:
                    ws.merge_cells('C'+str(i)+':E'+str(i))
            ws['D4'].alignment = self.align_right
            ws['E4'].alignment = self.align_center

            ws.column_dimensions['A'].width = 18
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 18

            ws['A1'] = "Обозначение объекта:"
            ws['A2'] = "Коэффициент светопропускания:"
            ws['A3'] = "Уровень светопропускания:"
            ws['A4'] = "Допустимое соответствие стекла:"
            ws['D4'] = "Статус:"
            ws['A5'] = "Целевая функция покрытия:"

            ws['C1'] = self.name_object_2
            ws['C2'] = self.koef_ltransm_str + ' %'
            ws['C3'] = self.level_ltransm_str
            ws['C4'] = self.Glass_compliance_str
            ws['E4'] = self.status_str
            ws['C5'] = self.Func_coverage.currentText()

            ws['A6'] = "Название нановещества"
            ws['B6'] = "Количественная оценка опасности"
            ws['C6'] = "Качественная оценка опасности"
            ws['D6'] = "Коэффициент неполноты анализа"
            ws['E6'] = "Оценка достоверности результатов"

            alf = {1: 'A', 2: 'B', 3: 'C', 4: 'D', 5: 'E'}
            for i in range(self.row_count_2):
                k = 1
                ws[alf[k] + str(i + 7)] = self.table_2.verticalHeaderItem(i).text()
                for j in range(4):
                    k += 1
                    ws[alf[k] + str(i + 7)] = self.table_2.item(i, j).text()

            ws.merge_cells('A'+str(end_table+1)+':B'+str(end_table+1))
            ws.merge_cells('C'+str(end_table+1)+':E'+str(end_table+1))
            ws.merge_cells('A'+str(end_table+2)+':B'+str(end_table+2))
            ws.merge_cells('C'+str(end_table+2)+':E'+str(end_table+2))
            ws['A'+str(end_table+1)].alignment = self.align_left
            ws['C'+str(end_table+1)].alignment = self.align_center
            ws['A'+str(end_table+2)].alignment = self.align_left
            ws['C'+str(end_table+2)].alignment = self.align_center

            ws['A'+str(end_table+1)] = "Общая степень потенциальной опасности исследуемого объекта:"
            ws['C'+str(end_table+1)] = self.Hazard_assessment_2_str
            ws['A'+str(end_table+2)] = "Категория безопасности:"
            ws['C'+str(end_table+2)] = self.Safety_category_2_str

            ws.merge_cells('A'+str(end_table+3)+':E'+str(end_table+3))
            ws.merge_cells('A'+str(end_table+4)+':E'+str(end_table+4))
            ws['A'+str(end_table+3)].alignment = self.align_center
            ws['A'+str(end_table+4)].alignment = self.align_center

            ws['A'+str(end_table+3)] = "Заключение о целесообразности применения продукта, "\
                                       "содержащего нановещества:"
            ws['A'+str(end_table+4)] = self.Conclusion_str
            ws.row_dimensions[end_table+4].height = 30

            wb.save(file_path)
            return
        except PermissionError:
            error_message = QtWidgets.QErrorMessage(self)
            error_message.setWindowTitle("Ошибка сохранения")
            error_message.showMessage("Сохранение не удалось, проверьте не\n"
                                      "открыт ли файл в другой программе")
        except FileNotFoundError:
            error_message = QtWidgets.QErrorMessage(self)
            error_message.setWindowTitle("Ошибка сохранения")
            error_message.showMessage("Вы не задали имя файла!")
        except Exception as ex:
            QMessageBox.question(self, 'Ошибка в программе',
                                 "Ошибка: " + str(type(ex)) + "; Значение: " + str(ex),
                                 QMessageBox.Ok)

    def AboutClicked(self):
        self.Icon.actualSize(QSize(150, 150), QIcon.Normal, QIcon.On)
        QMessageBox.about(self, "Программа ECO-NANO-TEST", ModuleNanoSubstances.about_info)

    def HelpClicked(self):
        self.window_help = InstructionsClass()
        self.window_help.show()


def main():
    app = QtWidgets.QApplication(sys.argv)
    window = MainClass()
    window.show()
    app.exec_()


if __name__ == "__main__":
    main()
