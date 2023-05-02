from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Side
from PyQt5 import QtWidgets, QtCore
from math import ceil


class DataStructure():
    def __init__(self, states, states_text, rangs, fi_values, p_rating, f_grade, incomp_ratio, f_grade_text,
                 incomp_ratio_text):
        self.states = states  # Оценка в баллах для установленного состояния признака
        self.states_text = states_text  # Выбранное состояние признака
        self.rangs = rangs  # Ранг
        self.fi_values = fi_values  # Значение «взвешивающей функции»
        self.p_rating = p_rating  # Частная опастность по функциональному блоку
        self.f_grade = f_grade  # Финальная оценка опастности вещества
        self.incomp_ratio = incomp_ratio  # Коэфициент достоверности результатов
        self.f_grade_text = f_grade_text  # Текстовая финальная оценка безопасности
        self.incomp_ratio_text = incomp_ratio_text  # Текстовая достоверность результатов


class ExcelReport(QtCore.QThread):
    reportErrSave = QtCore.pyqtSignal()

    def __init__(self, struc: DataStructure, name_material, save_path):
        super().__init__()
        self.data = struc
        self.path = save_path
        self.name_sheet = name_material

        self.font = Font(name='Times New Roman',
                         size=8,
                         bold=False,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000')

        self.fill = PatternFill(fill_type='solid',
                                start_color='c1c1c1',
                                end_color='c2c2c2')

        self.border = Border(left=Side(border_style='thin', color='FF000000'),
                             right=Side(border_style='thin', color='FF000000'),
                             top=Side(border_style='thin', color='FF000000'),
                             bottom=Side(border_style='thin', color='FF000000'),
                             diagonal=Side(border_style='thin', color='FF000000'),
                             diagonal_direction=0,
                             outline=Side(border_style='thin', color='FF000000'),
                             vertical=Side(border_style='thin', color='FF000000'),
                             horizontal=Side(border_style='thin', color='FF000000'))

        self.align_center = Alignment(horizontal='center',
                                      vertical='center',
                                      text_rotation=0,
                                      wrap_text=True,
                                      shrink_to_fit=False,
                                      indent=0)

        self.align_left = Alignment(horizontal='left',
                                    vertical='center',
                                    text_rotation=0,
                                    wrap_text=True,
                                    shrink_to_fit=False,
                                    indent=0)

        self.align_right = Alignment(horizontal='right',
                                     vertical='center',
                                     text_rotation=0,
                                     wrap_text=True,
                                     shrink_to_fit=False,
                                     indent=0)

        self.number_format = 'General'
        self.protection = Protection(locked=True, hidden=False)

    def run(self) -> None:
        try:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.title = self.name_sheet

            head = ["№ п/п", "Признаки", "Ранг", "Значение «взвешивающей функции»",
                    "Состояние признака", "Оценка в баллах", "Количественная оценка по блоку"]
            numeration = [1, 2, 3, 4, 5, 6, 7]
            block_name = ["Блок 1. Физические характеристики",
                          "Блок 2. Физико-химические свойства",
                          "Блок 3. Молекулярно-биологические свойства",
                          "Блок 4. Цитологические свойства",
                          "Блок 5. Физиологические свойства",
                          "Блок 6. Экологическая характеристика"]
            name_state = [["Минимальный размер частицы в одном из измерений",
                           "Формфактор (отношение максимального размера к минимальному)"],
                          ["Растворимость в воде",
                           "Растворимость в биологических жидкостях",
                           "Заряд",
                           "Адсорбционная ёмкость",
                           "Устойчивость к агрегации",
                           "Гидрофобность",
                           "Адгезия к поверхностям",
                           "Способность генерировать свободные радикалы"],
                          ["Взаимодействие с ДНК",
                           "Взаимодействие с белками",
                           "Взаимодействие с мембранами"],
                          ["Способность к накоплению в клетках",
                           "Трансформирующая активность",
                           "Влияние на протеомный и (или) метаболомный профиль",
                           "Токсичность для клеток"],
                          ["Проникновение через барьеры организма",
                           "Накопление в органах и тканях",
                           "Усиление проницаемости барьеров организма для посторонних токсикантов",
                           "Острая токсичность",
                           "Хроническая токсичность",
                           "Специфические и отдалённые эффекты токсичности (канцерогенный, "
                           "мутагенный, тератогенный, гонадотоксический, эмбриотоксический, "
                           "иммунотоксический. аллергенный)"],
                          ["Массовость производства в мире",
                           "Возможность экспонирования людей (категории населения)",
                           "Данные о накоплении в организмах",
                           "Данные о накоплении в объектах внешней среды (почвы, грунтовые воды, "
                           "донные отложения)"]
                          ]

            self.ws.append(head)
            self.ws.append(numeration)
            for i in range(len(block_name)):
                self.ws.append([block_name[i]])
                for j in range(len(self.data.states[i])):
                    line = [j + 1,
                            name_state[i][j],
                            self.data.rangs[i][j],
                            self.data.fi_values[i][j],
                            self.data.states_text[i][j],
                            self.data.states[i][j]]
                    self.ws.append(line)

            name_block_pos = [3, 6, 15, 19, 24, 31]
            for i in name_block_pos:
                self.ws.merge_cells('A{}:G{}'.format(i, i))

            p_rating_pos = [4, 7, 16, 20, 25, 32]
            for i in range(len(p_rating_pos)):
                self.ws['G{}'.format(p_rating_pos[i])] = ceil(self.data.p_rating[i] * 1000) / 1000
                self.ws.merge_cells('G{}:G{}'.format(p_rating_pos[i], p_rating_pos[i] + len(self.data.states[i]) - 1))

            for cellObj in self.ws['A1:G3']:
                for cell in cellObj:
                    self.ws[cell.coordinate].alignment = self.align_center

            for cellObj in self.ws['A4:A38']:
                for cell in cellObj:
                    self.ws[cell.coordinate].alignment = self.align_center

            for cellObj in self.ws['B4:B35']:
                for cell in cellObj:
                    self.ws[cell.coordinate].alignment = self.align_left

            for cellObj in self.ws['C4:D35']:
                for cell in cellObj:
                    self.ws[cell.coordinate].alignment = self.align_center

            for cellObj in self.ws['E4:E35']:
                for cell in cellObj:
                    self.ws[cell.coordinate].alignment = self.align_left

            for cellObj in self.ws['F4:G35']:
                for cell in cellObj:
                    self.ws[cell.coordinate].alignment = self.align_center

            for cellObj in self.ws['A1:G38']:
                for cell in cellObj:
                    self.ws[cell.coordinate].border = self.border
                    self.ws[cell.coordinate].font = self.font

            self.ws.column_dimensions['A'].width = 4.43
            self.ws.column_dimensions['B'].width = 22.57
            self.ws.column_dimensions['C'].width = 7.5
            self.ws.column_dimensions['D'].width = 11.71
            self.ws.column_dimensions['E'].width = 27.71
            self.ws.column_dimensions['F'].width = 6.43
            self.ws.column_dimensions['G'].width = 12.43

            self.ws.merge_cells('A36:C36')
            self.ws.merge_cells('E36:F36')
            self.ws['A36'] = "Коэффициент неполноты:"
            self.ws['A36'].alignment = self.align_right
            self.ws['D36'] = ceil(self.data.incomp_ratio * 1000) / 1000
            self.ws['D36'].alignment = self.align_center
            self.ws['E36'] = "Количественная оценка:"
            self.ws['E36'].alignment = self.align_right
            self.ws['G36'] = ceil(self.data.f_grade * 1000) / 1000
            self.ws['G36'].alignment = self.align_center

            self.ws.merge_cells('A37:C37')
            self.ws.merge_cells('D37:G37')
            self.ws['A37'] = "Качественная оценка:"
            self.ws['A37'].alignment = self.align_right
            self.ws['D37'] = self.data.f_grade_text
            self.ws['D37'].alignment = self.align_center

            self.ws.merge_cells('A38:C38')
            self.ws.merge_cells('D38:G38')
            self.ws['A38'] = "Оценка достоверности результатов:"
            self.ws['A38'].alignment = self.align_right
            self.ws['D38'] = self.data.incomp_ratio_text[0]
            self.ws['D38'].alignment = self.align_center

            self.wb.save(self.path)

        except PermissionError:
            self.reportErrSave.emit()
        except FileNotFoundError:
            QMessageBox.question(self, 'Ошибка сохранения',
                                 "Вы не задали имя файла!",
                                 QMessageBox.Ok)
        except Exception as ex:
            QMessageBox.question(self, 'Ошибка в программе',
                                 "Ошибка: " + str(type(ex)) + "; Значение: " + str(ex),
                                 QMessageBox.Ok)


